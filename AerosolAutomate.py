from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import pandas as pd
import os
import math


class AutomateAerosolRF():
    def __init__(self, mode):
        self.mode = mode
        self.cwd = os.getcwd()
        self.BasePath = os.path.join(self.cwd, "Data")
        self.names = ["one", "two", "three", "Top up", "Top down",
                      "Top dir", "Bottom up", "Bottom down", "Bottom dir"]
        self.filenames = self.extract_files()
        self.wb = Workbook()

        self.main_sheet = self.wb.active
        '''
        self.main_sheet.title = self.filenames["WA"].split(
            "/")[-1].split(".")[0].split("_")[0]
        '''
        self.main_sheet.title = "Flux Calculation"
        self.HandleDifferentMode()

    def extract_files(self):
        files = os.listdir(self.BasePath)
        filenames = {}
        for file in files:
            AerosolorNot = file.split(".")[0]
            if AerosolorNot == "na":
                filenames["NA"] = os.path.join(self.BasePath, file)
            elif AerosolorNot == "wa":
                filenames["WA"] = os.path.join(self.BasePath, file)
        return filenames

    def HandleDifferentMode(self):
        if self.mode == 'wa':
            self.df_wa = self.WithAerosolProcess()
        elif self.mode == 'na':
            self.df_na = self.WithoutAerosolProcess()
        elif self.mode == 'both':
            self.df_wa = self.WithAerosolProcess()
            self.df_na = self.WithoutAerosolProcess()

    def WithAerosolProcess(self):
        df = pd.read_csv(
            self.filenames["WA"], delim_whitespace=True, header=None, names=self.names)
        ws = self.wb.create_sheet("Sheet1")
        ws.title = self.filenames["WA"].split("/")[-1].split(".")[0]
        for i in range(3, len(self.names)):
            ws.cell(row=1, column=i+1).value = self.names[i]
        for ind in df.index:
            for i in range(len(self.names)):
                ws.cell(row=ind+2, column=i +
                        1).value = df.loc[ind, self.names[i]]
        return df

    def WithoutAerosolProcess(self):
        df = pd.read_csv(
            self.filenames["NA"], delim_whitespace=True, header=None, names=self.names)
        ws = self.wb.create_sheet("Sheet2")
        ws.title = self.filenames["NA"].split("/")[-1].split(".")[0]

        for i in range(3, len(self.names)):
            ws.cell(row=1, column=i+1).value = self.names[i]
        for ind in df.index:
            for i in range(len(self.names)):
                ws.cell(row=ind+2, column=i +
                        1).value = df.loc[ind, self.names[i]]
        return df

    def Calculate_fluxes(self, df):
        df = df[['Top up', 'Top down', 'Bottom up', 'Bottom down']]
        df['Flux @ TOA'] = df["Top up"]-df["Top down"]
        df['Flux @ Surface'] = df["Bottom up"]-df["Bottom down"]
        return df

    def Result(self):

        if self.mode == 'both' or self.mode == 'wa':
            self.df_wa = self.Calculate_fluxes(self.df_wa)

            self.main_sheet.cell(row=1, column=8).value = "WA"

            l = 0
            for col in self.df_wa.columns:
                self.main_sheet.cell(row=2, column=l+9).value = col
                l += 1

            l = 0
            for ind in self.df_wa.index:
                l = 0
                for col in self.df_wa.columns:
                    self.main_sheet.cell(
                        row=ind+3, column=l+9).value = self.df_wa.loc[ind, col]
                    l += 1

        if self.mode == 'both' or self.mode == 'na':
            #
            self.df_na = self.Calculate_fluxes(self.df_na)

            self.main_sheet.cell(row=1, column=1).value = "NA"
            l = 0

            for col in self.df_na.columns:
                self.main_sheet.cell(row=2, column=l+2).value = col
                l += 1
            l = 0
            for ind in self.df_na.index:
                l = 0
                for col in self.df_na.columns:
                    self.main_sheet.cell(
                        row=ind+3, column=l+2).value = self.df_na.loc[ind, col]
                    l += 1

        if self.mode == 'wa' or self.mode == 'na':
            self.wb.save(filename="filename.xlsx")
            print("Done")
            return

        TAS = []  # Total Aerosol at The surface
        TAA = []  # Total Aerosol at The TOA
        AF = []  # Atmospheric Flux
        for ind in self.df_wa.index:
            x = self.df_na.loc[ind, 'Flux @ Surface'] - \
                self.df_wa.loc[ind, 'Flux @ Surface']
            y = self.df_na.loc[ind, 'Flux @ TOA'] - \
                self.df_wa.loc[ind, 'Flux @ TOA']
            TAS.append(x)
            TAA.append(y)
            AF.append(y-x)

        col = ['Total Aerosol @ surface',
               'Total Aerosol @ TOA', 'Atmospheric Flux']
        for i in range(len(col)):
            self.main_sheet.cell(row=2, column=i+18).value = col[i]
        for i in range(len(TAS)):
            self.main_sheet.cell(row=i+3, column=18).value = TAS[i]
            self.main_sheet.cell(row=i+3, column=19).value = TAA[i]
            self.main_sheet.cell(row=i+3, column=20).value = AF[i]

        self.main_sheet.cell(row=27, column=17).value = "Avg"
        self.main_sheet.cell(row=27, column=18).value = sum(TAS)/len(TAS)
        self.main_sheet.cell(row=27, column=19).value = sum(TAA)/len(TAA)
        self.main_sheet.cell(row=27, column=20).value = sum(AF)/len(AF)

        for i in range(24):
            self.main_sheet.cell(row=i+40, column=1).value = i

        self.DrawChart(self.main_sheet, "A30", "TA @ Surface", 18)
        self.DrawChart(self.main_sheet, "I30", "TA @ TOA", 19)
        self.DrawChart(self.main_sheet, "R30", "Atomspheric Flux", 20)
        self.wb.save(filename="filename.xlsx")
        print("Done")

    def DrawChart(self, sheet, pos, title, refcol):
        xvalues = Reference(sheet, min_col=1, min_row=40, max_row=63)
        yvalues = Reference(sheet, min_col=refcol, min_row=3, max_row=26)
        series = Series(yvalues, xvalues)
        chart = ScatterChart()
        chart.series.append(series)
        chart.title = title
        chart.style = 13
        chart.x_axis.title = 'Time(UTC)'
        chart.y_axis.title = 'Radiative Forcing ( W/m*2 )'
        chart.legend = None
        sheet.add_chart(chart, pos)


#tt = AutomateAerosolRF('both')
#tt.Result()
